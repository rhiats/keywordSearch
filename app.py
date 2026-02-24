import streamlit as st
import pandas as pd
import re
from io import BytesIO

st.set_page_config(page_title="CSV Multi-Keyword Search", layout="wide")

st.title("🔎 High-Performance CSV Multi-Keyword Search")

uploaded_file = st.file_uploader("Upload your CSV file", type=["csv"])

if uploaded_file is not None:
    df = pd.read_csv(uploaded_file)

    st.success("File uploaded successfully!")
    st.dataframe(df.head())

    keyword_input = st.text_input(
        "Enter keywords (comma separated)",
        placeholder="apple, banana"
    )

    search_type = st.radio(
        "Search Logic",
        ["OR (any keyword matches)", "AND (all keywords must match)"]
    )

    column_option = st.selectbox(
        "Search in:",
        ["All Columns"] + list(df.columns)
    )

    if keyword_input:
        keywords = [k.strip() for k in keyword_input.split(",") if k.strip()]
        pattern = "|".join([re.escape(k) for k in keywords])
        
        df_str = df.astype(str)

        if column_option == "All Columns":
            contains = df_str.apply(
                lambda col: col.str.contains(pattern, case=False, regex=True)
            )
            if search_type.startswith("OR"):
                mask = contains.any(axis=1)
            else:
                mask = contains.all(axis=1)
        else:
            contains = df_str[column_option].str.contains(
                pattern, case=False, regex=True
            )
            if search_type.startswith("OR"):
                mask = contains
            else:
                mask = df_str[column_option].apply(
                    lambda x: all(
                        re.search(re.escape(k), str(x), re.IGNORECASE)
                        for k in keywords
                    )
                )

        filtered_df = df[mask].copy()

        st.write(f"### 🔍 Results ({len(filtered_df)} matches)")

        # Highlighting function
        def highlight_keywords(val):
            val_str = str(val)
            for keyword in keywords:
                val_str = re.sub(
                    f"({re.escape(keyword)})",
                    r"<mark>\1</mark>",
                    val_str,
                    flags=re.IGNORECASE
                )
            return val_str

        styled_df = filtered_df.style.format(
            lambda x: highlight_keywords(x)
        )

        st.write(styled_df.to_html(escape=False), unsafe_allow_html=True)

        # --------- Excel Export with Highlighting ---------
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            filtered_df.to_excel(writer, index=False, sheet_name="Results")
            workbook = writer.book
            worksheet = writer.sheets["Results"]

            from openpyxl.styles import PatternFill

            highlight_fill = PatternFill(start_color="FFFF00",
                                         end_color="FFFF00",
                                         fill_type="solid")

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell_value = str(cell.value)
                    if any(re.search(re.escape(k), cell_value, re.IGNORECASE) for k in keywords):
                        cell.fill = highlight_fill

        st.download_button(
            label="Download Highlighted Results (Excel)",
            data=output.getvalue(),
            file_name="filtered_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        filtered_df = df[mask].copy()

        # 🔥 SAVE TO SESSION STATE
        st.session_state["filtered_df"] = filtered_df

        st.write(f"### 🔍 Results ({len(filtered_df)} matches)")
        st.dataframe(filtered_df)